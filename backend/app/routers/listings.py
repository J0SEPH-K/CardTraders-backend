from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from typing import List, Optional
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from ..schemas.listings import Listing, ListingCreate
from ..db import get_db
from ..models.listing import Listing as ListingModel
from ..mongo import get_mongo_db, mongo_enabled, MONGODB_COLLECTION

router = APIRouter()

_DATA: List[Listing] = []

@router.get("/", response_model=List[Listing])
async def list_listings(db: Session = Depends(get_db), mdb=Depends(get_mongo_db)):
    # Fall back to in-memory if DB is not configured
    if mongo_enabled() and mdb is not None:
        docs = []
        async for d in mdb[MONGODB_COLLECTION].find({}).limit(1000):
            d["id"] = str(d.get("_id"))
            d.pop("_id", None)
            docs.append(Listing(**d))
        return docs
    try:
        rows = db.query(ListingModel).all()
        return [
            Listing(
                id=r.id,
                title=r.title,
                description=r.description,
                category=r.category, sport=r.sport, year=r.year, base=r.base,
                card_type=r.card_type, set_name=r.set_name, grade=r.grade,
                is_verified=r.is_verified, price=r.price,
            )
            for r in rows
        ]
    except Exception:
        return _DATA

@router.post("/", response_model=Listing)
async def create_listing(payload: ListingCreate, db: Session = Depends(get_db), mdb=Depends(get_mongo_db)):
    if mongo_enabled() and mdb is not None:
        doc = payload.model_dump()
        res = await mdb[MONGODB_COLLECTION].insert_one(doc)
        return Listing(id=str(res.inserted_id), **doc)
    try:
        model = ListingModel(
            **payload.model_dump(),
        )
        db.add(model)
        db.commit()
        db.refresh(model)
        return Listing(id=model.id, **payload.model_dump())
    except Exception:
        # in-memory fallback
        item = Listing(id=str(len(_DATA)+1), **payload.model_dump())
        _DATA.append(item)
        return item


@router.post("/upload-xlsx", response_model=int)
async def upload_xlsx(file: UploadFile = File(...), db: Session = Depends(get_db), mdb=Depends(get_mongo_db)):
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xlsm files are supported")

    content = await file.read()
    import io
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)

    ws = wb.active

    # Expected header mapping (pre-formatted table)
    # Adjust these to your template column titles
    header_map = {
        "title": ["Title", "제목"],
        "description": ["Description", "설명"],
        "category": ["Category", "카테고리"],
        "sport": ["Sport", "스포츠"],
        "year": ["Year", "년도"],
        "base": ["Base", "베이스"],
        "card_type": ["Card Type", "카드 유형"],
        "set_name": ["Set", "세트"],
        "grade": ["Grade", "등급"],
        "is_verified": ["Verified", "검증됨"],
        "price": ["Price", "가격"],
    }

    # Read header row
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Resolve column indices by matching any alias in header_map
    col_idx = {}
    for key, aliases in header_map.items():
        for i, name in enumerate(headers):
            if name in aliases:
                col_idx[key] = i
                break

    missing = [k for k in ("title", "category") if k not in col_idx]
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")

    created = 0
    # Iterate rows from row 2
    for row in ws.iter_rows(min_row=2):
        def get_val(key):
            i = col_idx.get(key)
            if i is None:
                return None
            v = row[i].value
            return v if v != "" else None

        # Normalize boolean-like values
        def to_bool(v):
            if isinstance(v, bool):
                return v
            if v is None:
                return False
            s = str(v).strip().lower()
            return s in {"1", "true", "yes", "y"}

        payload = ListingCreate(
            title=str(get_val("title") or "").strip(),
            description=(str(get_val("description")).strip() if get_val("description") is not None else None),
            category=str(get_val("category") or "").strip() or "pokemon",
            sport=(str(get_val("sport")).strip() if get_val("sport") is not None else None),
            year=(int(get_val("year")) if get_val("year") is not None else None),
            base=(str(get_val("base")).strip() if get_val("base") is not None else None),
            card_type=(str(get_val("card_type")).strip() if get_val("card_type") is not None else None),
            set_name=(str(get_val("set_name")).strip() if get_val("set_name") is not None else None),
            grade=(str(get_val("grade")).strip() if get_val("grade") is not None else None),
            is_verified=to_bool(get_val("is_verified")),
            price=(float(get_val("price")) if get_val("price") is not None else None),
        )

        # Persist
        doc = payload.model_dump()
        if mongo_enabled() and mdb is not None:
            try:
                await mdb[MONGODB_COLLECTION].insert_one(doc)
                created += 1
                continue
            except Exception:
                pass
        try:
            model = ListingModel(**doc)
            db.add(model)
            db.commit()
            db.refresh(model)
            created += 1
        except Exception:
            # memory fallback
            item = Listing(id=str(len(_DATA)+1), **doc)
            _DATA.append(item)
            created += 1

    return created
